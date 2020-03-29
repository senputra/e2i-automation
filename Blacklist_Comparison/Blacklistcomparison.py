import pandas as pd 
import openpyxl 
from openpyxl import Workbook
from openpyxl import load_workbook

#Create blacklist list 
def get_blacklistNRIC():
    #opening and reading  blacklist file 
    file = 'blacklist.xlsx'
    data = pd.ExcelFile(file)
    df = data.parse('blacklist')
    df.info 
    df.head()
    ps = openpyxl.load_workbook('blacklist.xlsx')
    sheet = ps ['blacklist']
    sheet.max_row
    #Creating a list of the blacklist NRIC 
    BNRIClist = []
    for row in range (3, sheet.max_row + 1):
        BNRIC = sheet ['B'+ str(row)].value 
        BNRIClist.append(BNRIC)
    return BNRIClist


#Next create a nric list from the compile file and cross compare
def get_consolidatedNRIC():
    #Opening and reading consolidated file 
    file1 = 'Consolidated_data.xlsx'
    data1 = pd.ExcelFile(file1)
    df1 = data1.parse('FINAL')
    df1.info 
    df1.head()
    lp = openpyxl.load_workbook('Consolidated_data.xlsx')
    sheet1 = lp ['FINAL']
    NRIClist = []
    for row in range(2, sheet1.max_row + 1):
        NRIC = sheet1['D'+ str(row)].value 
        NRIClist.append(NRIC)
    return NRIClist 

def create_matched_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched_Blacklisted_Data.xlsx"
    wb.save (filename = 'Matched_Blacklisted_Data.xlsx')



def get_updatedlist():
    mainlist = []
    wb = openpyxl.load_workbook('blacklist.xlsx')
    sheet = wb.active 
    max_row = sheet.max_row
    max_column = sheet.max_column
    for i in range(1,max_row+1):
        sublist = []
        for j in range (1,max_column+1):
            cell_obj = sheet.cell(row=i,column=j)
            #print (cell_obj.value, type(cell_obj.value))
            sublist.append(str(cell_obj.value))
        mainlist.append(sublist)
    updatedlist = []
    for a in overlaplist:
        counter = 0 
        while counter < len(mainlist):
            if a == mainlist[counter][1]:
                updatedlist.append(mainlist[counter])
            counter += 1
    return updatedlist
  


def matched_data():
    
    wb = load_workbook('Matched_Blacklisted_Data.xlsx')
    sheet = wb.active
    sheet['A1'] = 'S/N'
    sheet['B1'] = 'NRIC'
    sheet['C1'] = 'Start of suspension period for SSG Funding (dd/mm/yyyy)'
    sheet['D1'] = 'Start of suspension period for WSG Funding (dd/mm/yyyy)'
    sheet['D1'] = 'End of suspension period (dd/mm/yyyy)'
#     starterline = wb.max_row + 1 
    for row in O_updatedlist:
        sheet.append(row)
    wb.save('Matched_Blacklisted_Data.xlsx')
    

create_matched_excel() #Create empty excel file 
O_NRIClist = get_consolidatedNRIC()
O_BNRIClist = get_blacklistNRIC()

#Output a list of the overlap 
overlaplist = []
for nric in range(len(O_BNRIClist)):
    if O_BNRIClist[nric] in O_NRIClist:
        overlaplist.append(O_BNRIClist[nric])
O_updatedlist = []
convert = get_updatedlist()
for sublist in convert: 
    newtuple = tuple(sublist)
    O_updatedlist.append(newtuple)
    
matched_data()

print ('done')
    
    
    
    
    
    
    
    
    
#     for row in ws.iter_rows(ws.min_row,ws.max_row):
#         for cell in row: 
#             #print (type(cell.value), cell.value)
#             if type(cell.value) == str: 
#                 if cell.value in overlaplist: 
#                     rowlist.append(cell)
#     return rowlist 

# print (matched_excel())

                
                


    





