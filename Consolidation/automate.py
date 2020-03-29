from openpyxl import load_workbook
import pandas as pd

import sys
import argparse


def newDict(a):

    # Populating the initial dictionary with empty list
    newDict = {}
    for i in a:
        newDict[i] = []
    return newDict


def padDict(d, pad=""):
    length = len(d['Submission ID'])
    for key, value in d.items():
        if len(value) < length:
            d[key].append(pad)
    return d


def append_df_to_excel(filename, df=None, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """


    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    return writer, startrow
    # # write out the new sheet
    # df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # # save the workbook
    # writer.save()


def convert(inputFilePath='inputdata.xlsx', outputFilePath='outputData.xlsx', a=[
    'No',
    'Submission ID',
    'Name (As per NRIC)',
    'NRIC number',
    'Address (As per NRIC)',
    'Mailing address',
    'Contact number (Residential)',
    'Contact number (Mobile)',
    'Email',
    'SSIC',
    'NTUC Union Membership',
    'Apply Membership?﻿',
    'Preferred mode of payment',
    'Name (as per Bank record)',
    'Bank name',
    'Bank account number',
    'ACKNOWLEDGEMENT']):


    writer, startRow = append_df_to_excel(outputFilePath, sheet_name="Sheet1")
    rowIndex = startRow
    outDict = newDict(a)

    # Get Excel Book
    xls = pd.ExcelFile(inputFilePath)
    datasheetNames = xls.sheet_names  # To get the list of the name of datasheet

    # Loop every datasheet in the book
    for datasheetName in datasheetNames:
        # setup dictionary for the output file
        # print(outDict)
        outDict['No'].append(rowIndex)
        # outDict['Serial No'].append(0)
        rowIndex += 1

        df = xls.parse(datasheetName, header=None)
        convDict = df.to_dict()
        # populate the new dictionary with the data looped from input file
        for index, key in convDict[0].items():
            if outDict.get(str(convDict[0][index]).strip(), [-1]) != [-1]:
                outDict[str(convDict[0][index]).strip()
                        ].append(convDict[1][index])

        outDict = padDict(outDict)

    outDf = pd.DataFrame(data=outDict, index=None)
    if startRow == 0:
        outDf.to_excel(writer, "Sheet1", startrow=startRow, index=None)
    else:
        outDf.to_excel(writer, "Sheet1", startrow=startRow,
                       index=None, header=None)

    # save the workbook
    writer.save()
    return outDf


if __name__ == "__main__":
   
    def getOptions(args=sys.argv[1:]):
        parser = argparse.ArgumentParser(
            description="Excel automation script.")
        parser.add_argument("-i", "--input", help="Your input file.")
        parser.add_argument(
            "-o", "--output", help="Your destination output file. Default is 'outputdata.xlsx'")
        options = parser.parse_args(args)
        return options, parser

    options, parser = getOptions(sys.argv[1:])

    outputFilePath = '../outputdata.xlsx'
    inputFilePath = None

    inputFilePath = options.input if options.input is not None else None
    outputFilePath = options.output if options.output is not None else outputFilePath

    if not bool(outputFilePath and inputFilePath):
        print(parser.print_help())
        sys.exit()

    print("Processing {}. Output file can be found at {}".format(
        inputFilePath, outputFilePath))
    convert(inputFilePath, outputFilePath)
    print("Done")
